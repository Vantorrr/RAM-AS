import openpyxl
from openpyxl.styles import Font, PatternFill

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары для загрузки"

    # Define headers
    headers = [
        "Артикул (Part Number)*", 
        "Название товара*", 
        "Категория (из списка)*", 
        "Цена продажи (USD)*", 
        "Остаток (шт)*", 
        "Описание (необязательно)",
        "Фото (имя файла или ссылка)"
    ]

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        
        # Adjust column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25

    # Add a hint row
    hints = [
        "Пример: 68003701AC", 
        "Пример: Тормозной диск передний", 
        "Пример: Тормозная система", 
        "150.00", 
        "10", 
        "Оригинал MOPAR",
        "disk.jpg"
    ]
    ws.append(hints)

    # Save
    wb.save("template_ramus.xlsx")
    print("Template created: template_ramus.xlsx")

if __name__ == "__main__":
    create_template()



